#!/usr/bin/env python3
"""
Google Drive Handler: Recursive scan, dynamic folder creation, batch embedding.
Manages user document contexts with metadata tracking.
"""

import os
import json
import logging
import io
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
import hashlib

from google.oauth2 import service_account
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
import chromadb
from sentence_transformers import SentenceTransformer
import torch
import PyPDF2
from docx import Document

logger = logging.getLogger(__name__)

class DriveHandler:
    """Google Drive document handler with batch processing."""
    
    SUPPORTED_EXTENSIONS = {
        '.pdf': 'PDF', '.docx': 'Word', '.doc': 'Word',
        '.xlsx': 'Excel', '.xls': 'Excel', '.txt': 'Text', '.csv': 'CSV',
        '.jpg': 'Image', '.jpeg': 'Image', '.png': 'Image', '.bmp': 'Image'
    }
    
    def __init__(self, service_account_json: str):
        self.service_account_json = service_account_json
        self.drive_service = None
        self.encoder = None
        self.chroma_client = None
        self.processed_files_log = Path('drive_sync/processed_files.json')
        self.user_folders = {}
        
        self.authenticate()
        self._load_processed_files()
    
    def authenticate(self):
        """Authenticate with Google Drive."""
        try:
            if not os.path.exists(self.service_account_json):
                logger.error(f"Service account JSON not found: {self.service_account_json}")
                return False
            
            credentials = service_account.Credentials.from_service_account_file(
                self.service_account_json,
                scopes=['https://www.googleapis.com/auth/drive'] 
            )
            
            self.drive_service = build('drive', 'v3', credentials=credentials)
            self.chroma_client = chromadb.PersistentClient(path="./chroma_data")
            
            device = 'cuda' if torch.cuda.is_available() else 'cpu'
            self.encoder = SentenceTransformer('all-MiniLM-L6-v2')
            
            logger.info(f"Google Drive authenticated (device={device})")
            return True
        except Exception as e:
            logger.error(f"Authentication failed: {e}")
            return False
    
    def _load_processed_files(self):
        try:
            if self.processed_files_log.exists():
                with open(self.processed_files_log, 'r') as f:
                    self.processed_files = json.load(f)
            else:
                self.processed_files = {}
                Path('drive_sync').mkdir(exist_ok=True)
        except Exception as e:
            logger.error(f"Load processed files failed: {e}")
            self.processed_files = {}
    
    def _save_processed_files(self):
        try:
            Path('drive_sync').mkdir(exist_ok=True)
            with open(self.processed_files_log, 'w') as f:
                json.dump(self.processed_files, f, indent=2)
        except Exception as e:
            logger.error(f"Save processed files failed: {e}")
    
    def create_user_folder(self, user_id: str, user_email: Optional[str] = None, folder_name: str = None) -> Tuple[bool, str]:
        """Dynamically create folder for user in Drive and share it via Email."""
        try:
            if not self.drive_service:
                logger.error("Drive not authenticated")
                return False, ""
            
            folder_name = folder_name or f"SP-LineBot-{user_id}"
            
            query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
            results = self.drive_service.files().list(q=query, spaces='drive', fields="files(id, webViewLink)").execute()
            
            if results.get('files'):
                folder_id = results['files'][0]['id']
                folder_link = results['files'][0].get('webViewLink', '')
                logger.info(f"User folder already exists: {folder_id}")
            else:
                file_metadata = {
                    'name': folder_name,
                    'mimeType': 'application/vnd.google-apps.folder'
                }
                folder = self.drive_service.files().create(
                    body=file_metadata, 
                    fields='id, webViewLink'
                ).execute()
                
                folder_id = folder.get('id')
                folder_link = folder.get('webViewLink', '')
                logger.info(f"User folder created: {folder_id}")

            self.user_folders[user_id] = folder_id

            if user_email:
                try:
                    permission = {
                        'type': 'user',
                        'role': 'writer',
                        'emailAddress': user_email
                    }
                    self.drive_service.permissions().create(
                        fileId=folder_id,
                        body=permission,
                        sendNotificationEmail=True
                    ).execute()
                    logger.info(f"Successfully sent folder invite to {user_email}")
                except Exception as e:
                    logger.error(f"Failed to share folder with {user_email}: {e}")
                    return False, folder_link
            
            return True, folder_link
        except Exception as e:
            logger.error(f"Create user folder failed: {e}")
            return False, ""

    def scan_user_folder(self, user_id: str, folder_id: Optional[str] = None) -> List[Dict[str, Any]]:
        try:
            if not self.drive_service:
                return []
            
            folder_id = folder_id or self.user_folders.get(user_id)
            if not folder_id:
                return []
            
            documents = []
            self._scan_folder_recursive(folder_id, documents, user_id)
            return documents
        except Exception as e:
            logger.error(f"Folder scan failed: {e}")
            return []
    
    def _scan_folder_recursive(self, folder_id: str, documents: List[Dict], user_id: str, depth: int = 0):
        try:
            if depth > 5: return
            
            query = f"'{folder_id}' in parents and trashed=false"
            results = self.drive_service.files().list(
                q=query, spaces='drive', pageSize=100,
                fields='files(id, name, mimeType, size, modifiedTime)'
            ).execute()
            
            for file in results.get('files', []):
                file_id, name, mime_type = file['id'], file['name'], file['mimeType']
                size, mod_time = int(file.get('size', 0)), file.get('modifiedTime')
                
                if mime_type == 'application/vnd.google-apps.folder':
                    self._scan_folder_recursive(file_id, documents, user_id, depth + 1)
                else:
                    ext = Path(name).suffix.lower()
                    if ext in self.SUPPORTED_EXTENSIONS:
                        file_hash = hashlib.md5(f"{file_id}_{mod_time}".encode()).hexdigest()
                        if file_hash in self.processed_files: continue
                        
                        documents.append({
                            'id': file_id, 'name': name, 'type': self.SUPPORTED_EXTENSIONS[ext],
                            'size_bytes': size, 'modified': mod_time, 'hash': file_hash, 'user_id': user_id
                        })
        except Exception as e:
            logger.error(f"Recursive scan failed: {e}")

    # ========================================================================
    # UPDATED: BATCH EMBEDDING WITH SMART CSV ROUTING
    # ========================================================================
    
    def batch_embed_documents(self, documents: List[Dict[str, Any]], user_id: str) -> int:
        if not self.chroma_client or not self.encoder: return 0
        embedded_count = 0
        
        from drive_scanner import parse_dense_inventory_csv
        
        try:
            collection = self.chroma_client.get_or_create_collection(name=f"drive_user_{user_id}")
            for doc in documents:
                try:
                    ext = Path(doc['name']).suffix.lower()
                    request = self.drive_service.files().get_media(fileId=doc['id'])
                    file_stream = io.BytesIO()
                    downloader = MediaIoBaseDownload(file_stream, request)
                    done = False
                    while done is False:
                        status, done = downloader.next_chunk()
                    
                    temp_path = f"temp_{doc['id']}{ext}"
                    with open(temp_path, 'wb') as f:
                        f.write(file_stream.getbuffer())
                    
                    chunks = []
                    if ext == '.csv':
                        logger.info(f"Using semantic unroller for {doc['name']}")
                        chunks = parse_dense_inventory_csv(temp_path)
                    else:
                        text = ""
                        if ext == '.pdf': text = self._extract_pdf(temp_path)
                        elif ext in ['.docx', '.doc']: text = self._extract_docx(temp_path)
                        elif ext == '.txt':
                            with open(temp_path, 'r', encoding='utf-8', errors='ignore') as f: text = f.read()
                        elif ext in ['.xlsx', '.xls']: 
                            text = self._extract_excel(temp_path)
                        if text:
                            chunks = self._chunk_text(text, chunk_size=500)
                    
                    if os.path.exists(temp_path): os.remove(temp_path)
                    if not chunks: continue
                    
                    for i, chunk in enumerate(chunks):
                        embedding = self.encoder.encode(chunk, convert_to_tensor=False)
                        collection.add(
                            ids=[f"{user_id}_{doc['id']}_{i}"],
                            embeddings=[embedding.tolist()],
                            metadatas=[{'user_id': user_id, 'file_id': doc['id'], 'file_name': doc['name'], 'chunk_idx': i}],
                            documents=[chunk]
                        )
                        embedded_count += 1
                    
                    self.processed_files[doc['hash']] = {'file_id': doc['id'], 'name': doc['name']}
                except Exception as e: logger.error(f"Embedding failed: {e}")
            
            self._save_processed_files()
            return embedded_count
        except Exception as e:
            logger.error(f"Batch embedding failed: {e}")
            return 0
            
    # Legacy extraction methods for non-CSV files
    def _extract_pdf(self, file_path: str) -> str:
        try:
            text = []
            with open(file_path, 'rb') as f:
                for page in PyPDF2.PdfReader(f).pages: text.append(page.extract_text())
            return '\n'.join(text)
        except: return ""
    
    def _extract_docx(self, file_path: str) -> str:
        try: return '\n'.join([p.text for p in Document(file_path).paragraphs])
        except: return ""
    
    def _extract_excel(self, file_path: str) -> str:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            text = []
            for sheet in wb.sheetnames:
                for row in wb[sheet].iter_rows(values_only=True):
                    text.append(' | '.join(str(cell or '') for cell in row))
            return '\n'.join(text)
        except: return ""
        
    def _chunk_text(self, text: str, chunk_size: int = 500) -> List[str]:
        words = text.split()
        return [' '.join(words[i:i+chunk_size]) for i in range(0, len(words), chunk_size)] if words else [text]

    def fetch_user_drive_context(self, user_id: str, top_k: int = 3) -> Dict[str, Any]:
        if not self.chroma_client: 
            return {}
        try:
            count = self.chroma_client.get_or_create_collection(name=f"drive_user_{user_id}").count()
            return {'user_id': user_id, 'document_count': count, 'status': 'ready' if count > 0 else 'empty'}
        except Exception as e: 
            return {'error': str(e)}